from download import write_to_txt_file, write_to_xlsx_file
from openpyxl import load_workbook



def extract_xlsx_files(path):

    #extrait le workbook puis le worksheet
    wb = load_workbook(path, data_only=True)
    ws = wb['List of channels']

    all_col = list(ws.columns)


    for cell in all_col[6]:
        content = cell.value
        try:
            write_to_txt_file('tv channels', content);
        except Exception as e:
            write_to_txt_file('errors.txt', f"erreur pour {content}: {e}")


def get_all_specchar(path):
    wb = load_workbook(path, data_only=True)
    ws = wb['List of channels']

    all_col = list(ws.columns)

    res = []
    for cell in all_col[0]:
        content = cell.value
        try:
            i = 0
            n = len(content)
            while i < n-3:
                char = content[i]
                if not (char.isalpha() or char.isnumeric()) and char != " ":
                   res.append([content])
                   i = n
                i += 1
        except Exception as e:
            write_to_txt_file("errors.txt", str(e))
    return res


if __name__ == '__main__':
   write_to_xlsx_file("channels_names/channel_list_charspec.xlsx", get_all_specchar("channels_names/20240318 list of tv channels.xlsx"))