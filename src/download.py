#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Module de téléchargement pour le TV Logo Downloader
Ce module contient les fonctions de téléchargement et de gestion des noms de chaînes.
"""

import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

# Dictionnaire pour la conversion des caractères spéciaux
SPECIAL_CHAR_MAPPING = {
    "+": "Plus", "&": "And", "'": "", "_": "", 
    "(": "", "/": "", ")": "", "-": "", 
    ".": "", ";": "", " ": "", "*": "", 
    "!": "", "´": ""
}

# Utilitaires de fichiers
def ensure_directory_exists(path):
    """Crée le répertoire s'il n'existe pas."""
    if not os.path.exists(path):
        os.makedirs(path)
        print(f"Répertoire créé: {path}")

def write_to_txt_file(file_path, content):
    """Écrit du contenu dans un fichier texte."""
    directory = os.path.dirname(file_path)
    ensure_directory_exists(directory)
    
    with open(file_path, 'a', encoding='utf-8') as file:
        file.write(content)
        file.write('\n')

def write_to_xlsx_file(file_path, content, column_name='channel_name'):
    """Écrit du contenu dans un fichier Excel."""
    directory = os.path.dirname(file_path)
    ensure_directory_exists(directory)
    
    df = pd.DataFrame(content, columns=[column_name])
    df.to_excel(file_path, index=False)
    print(f"Données écrites dans {file_path}")

# Téléchargement des logos
def download_logo(channel_name, output_dir):
    """
    Télécharge le logo d'une chaîne de télévision depuis Google Images.
    
    Args:
        channel_name (str): Nom de la chaîne
        output_dir (str): Répertoire de sortie où sauvegarder l'image
    
    Returns:
        bool: True si le téléchargement a réussi, False sinon
    """
    # Prépare la requête de recherche Google
    if len(channel_name) > 2 and channel_name[-3] == '.':
        search_term = f"{channel_name[:-3]}+logo"
    else:
        search_term = f"{channel_name}+logo"
    
    search_term = search_term.replace(' ', '+')
    google_search_url = f"https://www.google.com/search?hl=en&tbm=isch&q={search_term}"
    
    try:
        # Récupération de la page de résultats
        response = requests.get(google_search_url)
        response.raise_for_status()  # Vérifie si la requête a réussi
        
        # Analyse de la page HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        images = soup.findAll('img')
        
        # Sélection de l'image
        current_image_url = None
        for image in images:
            if 'src' in image.attrs:
                # Préfère les images de Wikipedia pour leur qualité
                if 'wikipedia' in image['src']:
                    current_image_url = image['src']
                    break
        
        # Si aucune image Wikipedia n'a été trouvée, prend la première image (en ignorant le logo Google)
        if not current_image_url and len(images) > 1:
            current_image_url = images[1]['src']
        
        if not current_image_url:
            print(f"Aucune image trouvée pour {channel_name}")
            return False
        
        # Téléchargement de l'image
        image_response = requests.get(current_image_url)
        image_response.raise_for_status()
        
        # Enregistrement de l'image
        output_path = f'{output_dir}/{channel_name}.png'
        with open(output_path, 'wb') as file:
            file.write(image_response.content)
        
        return True
    
    except requests.RequestException as e:
        print(f"Erreur lors de la requête pour {channel_name}: {e}")
        return False
    except Exception as e:
        print(f"Erreur inattendue pour {channel_name}: {e}")
        return False

# Extraction des noms de chaînes depuis Excel
def extract_channel_names(excel_path, worksheet_name='List of channels'):
    """
    Extrait la liste des noms de chaînes à partir d'un fichier Excel.
    
    Args:
        excel_path (str): Chemin vers le fichier Excel
        worksheet_name (str): Nom de la feuille contenant la liste des chaînes
    
    Returns:
        list: Liste des noms de chaînes
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[worksheet_name]
        
        all_columns = list(ws.columns)
        channel_names = []
        
        # Supposons que les noms de chaînes sont dans la première colonne
        for cell in all_columns[0]:
            content = cell.value
            if content:
                channel_names.append(content)
        
        return channel_names
    
    except Exception as e:
        write_to_txt_file('logs/errors.txt', f"Erreur lors de l'extraction des noms de chaînes: {e}")
        return []

def extract_channels_with_special_chars(excel_path, worksheet_name='List of channels'):
    """
    Extrait la liste des noms de chaînes contenant des caractères spéciaux.
    
    Args:
        excel_path (str): Chemin vers le fichier Excel
        worksheet_name (str): Nom de la feuille contenant la liste des chaînes
    
    Returns:
        list: Liste des noms de chaînes contenant des caractères spéciaux
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[worksheet_name]
        
        all_columns = list(ws.columns)
        channels_with_special_chars = []
        
        for cell in all_columns[0]:
            content = cell.value
            if not content:
                continue
                
            for char in content:
                if not (char.isalpha() or char.isnumeric() or char == " "):
                    channels_with_special_chars.append(content)
                    break
        
        return channels_with_special_chars
    
    except Exception as e:
        write_to_txt_file('logs/errors.txt', f"Erreur lors de l'extraction des chaînes avec caractères spéciaux: {e}")
        return []

# Traitement des noms de chaînes
def normalize_channel_name(channel_name):
    """
    Normalise un nom de chaîne en remplaçant les caractères spéciaux.
    
    Args:
        channel_name (str): Nom de chaîne original
    
    Returns:
        str: Nom de chaîne normalisé
    """
    normalized_name = ""
    
    for char in channel_name:
        if not (char.isalpha() or char.isnumeric()):
            normalized_name += SPECIAL_CHAR_MAPPING.get(char, "")
        else:
            normalized_name += char
    
    return normalized_name

def normalize_channel_list(channel_list):
    """
    Normalise une liste de noms de chaînes.
    
    Args:
        channel_list (list): Liste des noms de chaînes originaux
    
    Returns:
        list: Liste des noms de chaînes normalisés
    """
    return [normalize_channel_name(channel) for channel in channel_list]

def download_all_logos(channel_list_path, output_dir):
    """
    Télécharge les logos pour toutes les chaînes listées dans un fichier.
    
    Args:
        channel_list_path (str): Chemin vers le fichier contenant la liste des chaînes
        output_dir (str): Répertoire où sauvegarder les logos
    """
    ensure_directory_exists(output_dir)
    
    try:
        with open(channel_list_path, 'r', encoding='utf-8') as file:
            channels = [line.strip() for line in file.readlines()]
    except UnicodeDecodeError:
        # Essai avec un autre encodage si utf-8 échoue
        with open(channel_list_path, 'r', encoding='latin-1') as file:
            channels = [line.strip() for line in file.readlines()]
    
    channels.sort()
    total_channels = len(channels)
    
    for index, channel in enumerate(channels):
        try:
            success = download_logo(channel, output_dir)
            progress = round((index + 1) / total_channels * 100, 2)
            
            if success:
                print(f"Logo téléchargé pour {channel} ({progress}%)")
            else:
                print(f"Échec du téléchargement pour {channel} ({progress}%)")
                write_to_txt_file('logs/errors.txt', f"Échec du téléchargement pour {channel}")
        
        except Exception as e:
            write_to_txt_file('logs/errors.txt', f"Erreur lors du téléchargement du logo pour {channel}: {e}")

def prepare_normalized_channel_list(excel_path, output_file):
    """
    Prépare une liste normalisée des chaînes à partir d'un fichier Excel et l'enregistre.
    
    Args:
        excel_path (str): Chemin vers le fichier Excel source
        output_file (str): Chemin vers le fichier de sortie
    """
    # Extraire les chaînes avec caractères spéciaux
    channels_with_special_chars = extract_channels_with_special_chars(excel_path)
    
    # Normaliser les noms de chaînes
    normalized_channels = normalize_channel_list(channels_with_special_chars)
    
    # Écrire dans un fichier Excel
    write_to_xlsx_file(output_file, normalized_channels)
    print(f"Liste des chaînes normalisées enregistrée dans {output_file}")