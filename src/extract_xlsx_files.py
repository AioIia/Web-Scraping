from download import write_to_txt_file
from openpyxl import load_workbook


def get_channel_name(path):

    #extrait le workbook puis le worksheet
    wb = load_workbook(path, data_only=True)
    ws = wb['List of channels']

    all_col = list(ws.columns)
    res = []

    for cell in all_col[0]:
        content = cell.value
        try:
            if content:
                res.append(content)
        except Exception as e:
            write_to_txt_file('errors.txt', f"erreur pour {content}: {e}")
    return res


def get_all_specchar(path):
    wb = load_workbook(path, data_only=True)
    ws = wb['List of channels']

    all_col = list(ws.columns)

    res = []
    for cell in all_col[0]:
        content = cell.value
        try:
            i = 0
            n = len(content)
            while i < n:
                char = content[i]
                if not (char.isalpha() or char.isnumeric()) and char != " ":
                   res.append(content)
                   i = n
                i += 1
        except Exception as e:
            write_to_txt_file("errors.txt", str(e))
    return res

if __name__ == '__main__':
    print(get_channel_name("data/channels_names/20240318 list of tv channels.xlsx"))