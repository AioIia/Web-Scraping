#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Module de traitement pour le TV Logo Downloader
Ce module contient les fonctions de traitement des images.
"""

import os
import time
import logging
from PIL import Image
import requests
from bs4 import BeautifulSoup
import openpyxl
from tqdm import tqdm

# Configuration du logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/process.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Dictionnaire de codes pays EPG
country_codes = dict([
    ("france", "fr"), ("germany", "de"), ("united states", "us"), ("usa", "us"), ("italy", "it"),
    ("united kingdom", "uk"), ("canada", "ca"), ("spain", "es"), ("australia", "au"), ("japan", "jp"),
    ("china", "cn"), ("russia", "ru"), ("brazil", "br"), ("india", "in"), ("netherlands", "nl"),
    ("sweden", "se"), ("mexico", "mx"), ("south korea", "kr"), ("south africa", "za"), ("argentina", "ar"),
    ("switzerland", "ch"), ("norway", "no"), ("denmark", "dk"), ("finland", "fi"), ("austria", "at"),
    ("belgium", "be"), ("greece", "gr"), ("portugal", "pt"), ("poland", "pl"), ("turkey", "tr"),
    ("saudi arabia", "sa"), ("uae", "ae"), ("singapore", "sg"), ("malaysia", "my"), ("indonesia", "id"),
    ("thailand", "th"), ("new zealand", "nz"), ("ireland", "ie"), ("hungary", "hu"), ("czech republic", "cz"),
    ("romania", "ro"), ("egypt", "eg"), ("nigeria", "ng"), ("kenya", "ke"), ("israel", "il"),
    ("chile", "cl"), ("colombia", "co"), ("peru", "pe"), ("venezuela", "ve"), ("pakistan", "pk"),
    ("bangladesh", "bd"), ("philippines", "ph"), ("vietnam", "vn"), ("ukraine", "ua"), ("kazakhstan", "kz"),
    ("algeria", "dz"), ("morocco", "ma"), ("ethiopia", "et"), ("ghana", "gh"), ("uganda", "ug"),
    ("zimbabwe", "zw"), ("mozambique", "mz"), ("angola", "ao"), ("madagascar", "mg"), ("cameroon", "cm"),
    ("ivory coast", "ci"), ("niger", "ne"), ("mali", "ml"), ("senegal", "sn"), ("tunisia", "tn"),
    ("libya", "ly"), ("jordan", "jo"), ("lebanon", "lb"), ("syria", "sy"), ("iraq", "iq"),
    ("iran", "ir"), ("afghanistan", "af"), ("nepal", "np"), ("sri lanka", "lk"), ("myanmar", "mm"),
    ("cambodia", "kh"), ("laos", "la"), ("mongolia", "mn"), ("taiwan", "tw"), ("hong kong", "hk"),
    ("macau", "mo"), ("north korea", "kp"), ("cuba", "cu"), ("costa rica", "cr"), ("panama", "pa"),
    ("honduras", "hn"), ("guatemala", "gt"), ("el salvador", "sv"), ("nicaragua", "ni"), ("ecuador", "ec"),
    ("bolivia", "bo"), ("paraguay", "py"), ("uruguay", "uy"), ("guyana", "gy"), ("suriname", "sr"),
    ("french guiana", "gf"), ("belarus", "by"), ("moldova", "md"), ("armenia", "am"), ("georgia", "ge"),
    ("azerbaijan", "az"), ("kyrgyzstan", "kg"), ("tajikistan", "tj"), ("turkmenistan", "tm"), ("uzbekistan", "uz"),
    ("lithuania", "lt"), ("latvia", "lv"), ("estonia", "ee"), ("slovakia", "sk"), ("slovenia", "si"),
    ("croatia", "hr"), ("bosnia and herzegovina", "ba"), ("serbia", "rs"), ("montenegro", "me"), ("albania", "al"),
    ("north macedonia", "mk"), ("bulgaria", "bg"), ("cyprus", "cy"), ("malta", "mt"), ("iceland", "is"),
    ("luxembourg", "lu"), ("monaco", "mc"), ("andorra", "ad"), ("liechtenstein", "li"), ("san marino", "sm"),
    ("vatican city", "va"), ("kosovo", "xk"), ("puerto rico", "pr"), ("greenland", "gl"), ("faroe islands", "fo"),
    ("fiji", "fj"), ("papua new guinea", "pg"), ("solomon islands", "sb"), ("vanuatu", "vu"), ("new caledonia", "nc"),
    ("french polynesia", "pf"), ("western sahara", "eh"), ("sudan", "sd"), ("eritrea", "er"), ("djibouti", "dj"),
    ("somalia", "so"), ("mauritania", "mr"), ("gambia", "gm"), ("guinea-bissau", "gw"), ("guinea", "gn"),
    ("sierra leone", "sl"), ("liberia", "lr"), ("burkina faso", "bf"), ("togo", "tg"), ("benin", "bj"),
    ("mauritius", "mu"), ("seychelles", "sc"), ("comoros", "km"), ("botswana", "bw"), ("lesotho", "ls"),
    ("eswatini", "sz"), ("namibia", "na"), ("malawi", "mw"), ("zambia", "zm"), ("rwanda", "rw"),
    ("burundi", "bi"), ("chad", "td"), ("central african republic", "cf"), ("democratic republic of the congo", "cd"),
    ("republic of the congo", "cg"), ("gabon", "ga"), ("equatorial guinea", "gq"), ("sao tome and principe", "st"),
    ("cape verde", "cv"), ("maldives", "mv"), ("bhutan", "bt"), ("timor-leste", "tl"), ("brunei", "bn"),
    ("palau", "pw"), ("micronesia", "fm"), ("marshall islands", "mh"), ("kiribati", "ki"), ("tuvalu", "tv"),
    ("nauru", "nr"), ("samoa", "ws"), ("tonga", "to"), ("cook islands", "ck"), ("niue", "nu"),
    ("tokelau", "tk"), ("wallis and futuna", "wf"), ("american samoa", "as"), ("guam", "gu"),
    ("northern mariana islands", "mp"), ("palestine", "ps"), ("curaçao", "cw"), ("aruba", "aw"), ("bonaire", "bq"),
    ("sint maarten", "sx"), ("saint martin", "mf"), ("saint barthélemy", "bl"), ("saint pierre and miquelon", "pm"),
    ("montserrat", "ms"), ("british virgin islands", "vg"), ("us virgin islands", "vi"), ("anguilla", "ai"),
    ("saint kitts and nevis", "kn"), ("turks and caicos islands", "tc"), ("bahamas", "bs"), ("cayman islands", "ky"),
    ("bermuda", "bm"), ("saint vincent and the grenadines", "vc"), ("grenada", "gd"), ("barbados", "bb"),
    ("antigua and barbuda", "ag"), ("dominica", "dm"), ("saint lucia", "lc"), ("trinidad and tobago", "tt"),
    ("jamaica", "jm"), ("haiti", "ht"), ("dominican republic", "do"), ("martinique", "mq"), ("guadeloupe", "gp"),
    ("saint helena", "sh"), ("ascension island", "ac"), ("tristan da cunha", "ta"), ("falkland islands", "fk"),
    ("south georgia and the south sandwich islands", "gs"), ("bouvet island", "bv"),
    ("heard island and mcdonald islands", "hm"), ("french southern and antarctic lands", "tf"),
    ("svalbard and jan mayen", "sj"), ("antarctica", "aq")
])


def detect_country(channel_name, max_retries=3, retry_delay=2):
    """
    Recherche le pays d'une chaîne en ligne (via Google).
    
    Args:
        channel_name (str): Nom de la chaîne à rechercher
        max_retries (int): Nombre maximal de tentatives en cas d'échec
        retry_delay (int): Délai en secondes entre chaque tentative
        
    Returns:
        str: Code pays à deux lettres ou "??" si non trouvé
    """
    query = f"{channel_name} country"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    for attempt in range(max_retries):
        try:
            res = requests.get(f"https://www.google.com/search?hl=en&tbm=isch&q={query}", headers=headers)
            if res.status_code == 200:
                soup = BeautifulSoup(res.text, "html.parser")
                text = soup.get_text().lower()

                for country in country_codes:
                    if country in text:
                        return country_codes[country]
                
                # Vérifier des mots-clés supplémentaires pouvant indiquer le pays
                keywords = {
                    "canadian": "ca", "american": "us", "british": "uk", "french": "fr",
                    "german": "de", "spanish": "es", "italian": "it", "australian": "au"
                }
                
                for keyword, code in keywords.items():
                    if keyword in text:
                        return code
                        
                break  # Sortir si la requête a réussi mais aucun pays n'a été trouvé
            else:
                logger.warning(f"La requête pour {channel_name} a retourné un statut {res.status_code}")
        
        except Exception as e:
            logger.error(f"Erreur lors de la recherche pour {channel_name} (tentative {attempt+1}/{max_retries}): {e}")
            
            if attempt < max_retries - 1:
                logger.info(f"Nouvelle tentative dans {retry_delay} secondes...")
                time.sleep(retry_delay)
    
    return "??"  # Code inconnu

def create_excel(fichier_txt, fichier_excel="data/channels.xlsx"):
    """
    Crée un fichier Excel à partir d'un fichier texte contenant des noms de chaînes.
    
    Args:
        fichier_txt (str): Chemin vers le fichier texte source
        fichier_excel (str): Chemin vers le fichier Excel de sortie
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chaînes et Pays"

        # Ajouter l'en-tête avec formatage
        ws.append(["Chaîne", "Pays (epg.best)"])
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Compteurs pour le suivi
        total_lines = sum(1 for _ in open(fichier_txt, "r", encoding="utf-8"))
        processed = 0
        successful = 0
        unknown = 0

        # Création de la barre de progression
        with open(fichier_txt, "r", encoding="utf-8") as f:
            for ligne in tqdm(f, total=total_lines, desc="Création Excel", unit="chaînes"):
                mot = ligne.strip()
                if not mot:
                    continue

                processed += 1
                
                code = detect_country(mot)
                if code != "??":
                    successful += 1
                else:
                    unknown += 1
                
                ws.append([mot, code if code != "??" else "inconnu"])
                
                # Pause pour éviter d'être bloqué par les services de recherche
                if processed % 10 == 0:
                    time.sleep(2)

        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Sauvegarder le fichier
        wb.save(fichier_excel)
    
    except Exception as e:
        logger.error(f"Erreur lors de la création du fichier Excel : {e}")
        raise

def ensure_directory_exists(path):
    """Crée le répertoire s'il n'existe pas."""
    if not os.path.exists(path):
        os.makedirs(path)
        logger.info(f"Répertoire créé: {path}")

def write_to_txt_file(file_path, content):
    """Écrit du contenu dans un fichier texte."""
    directory = os.path.dirname(file_path)
    ensure_directory_exists(directory)
    
    with open(file_path, 'a', encoding='utf-8') as file:
        file.write(content)
        file.write('\n')

def process_images(directory, min_size=320, scale_factor=0.85):
    """
    Traite toutes les images dans un répertoire pour les uniformiser:
    - Redimensionne les petites images
    - Place les images sur un fond blanc carré
    
    Args:
        directory (str): Répertoire contenant les images
        min_size (int): Taille minimale des images
        scale_factor (float): Facteur d'échelle pour la taille des images
    """
    if not os.path.exists(directory):
        logger.error(f"Le répertoire {directory} n'existe pas.")
        return
    
    image_files = [f for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f)) 
                   and f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
    total_images = len(image_files)
    
    if total_images == 0:
        logger.warning(f"Aucune image trouvée dans {directory}")
        return
        
    processed = 0
    errors = 0
    
    # Barre de progression tqdm
    for image_filename in tqdm(image_files, desc="Traitement des images", unit="image"):
        try:
            image_path = os.path.join(directory, image_filename)
            
            img = Image.open(image_path)
            
            # Convertir en RGBA pour gérer la transparence
            if img.mode != 'RGBA':
                img = img.convert('RGBA')
            
            # Déterminer la taille maximale
            size_taken = max(img.size)
            
            # Redimensionner si l'image est trop petite
            if size_taken < min_size:
                if size_taken == img.width:
                    new_width = int(min_size * scale_factor)
                    new_height = int(img.height * min_size / img.width * scale_factor)
                    img = img.resize((new_width, new_height), Image.LANCZOS)
                else:
                    new_width = int(img.width * min_size / img.height * scale_factor)
                    new_height = int(min_size * scale_factor)
                    img = img.resize((new_width, new_height), Image.LANCZOS)
                size_taken = min_size
            
            # Créer un canevas blanc carré
            canvas = Image.new('RGBA', (size_taken, size_taken), (255, 255, 255, 255))
            
            # Calcul de la position pour centrer l'image
            position = (
                int((size_taken - img.width) / 2),
                int((size_taken - img.height) / 2)
            )
            
            # Coller l'image sur le canevas et sauvegarder
            canvas.paste(img, position, img if img.mode == 'RGBA' else None)
            
            # Convertir en RGB pour sauvegarder en PNG
            if canvas.mode == 'RGBA':
                canvas = canvas.convert('RGB')
                
            canvas.save(image_path)
            processed += 1
            
        except Exception as e:
            logger.error(f"Erreur de traitement pour {image_filename}: {e}")
            write_to_txt_file('logs/errors.txt', f"Erreur de traitement pour {image_filename}: {e}")
            errors += 1